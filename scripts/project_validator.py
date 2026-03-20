#!/usr/bin/env python

import re
import sys
import warnings
from argparse import ArgumentParser
from io import BytesIO

import psycopg2
import yaml
from genologics.config import BASEURI, PASSWORD, USERNAME
from genologics.entities import File, Project
from genologics.lims import Lims
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

with open("/opt/gls/clarity/users/glsai/config/genosqlrc.yaml") as f:
    config = yaml.safe_load(f)

DESC = """EPP used to validate a project before it is released, including checking sample name format, index format and index distance in library pool.
"""

# Pre-compile regexes in global scope:
NGISAMPLE_PAT = re.compile("P[0-9]+_[0-9]+")
IDX_PAT = re.compile("([ATCG]{4,}N*)-?([ATCG]*)")
VALIDBASES_PAT = re.compile(r"^[ATCGN\-]+$")
TENX_SINGLE_PAT = re.compile("SI-(?:GA|NA)-[A-H][1-9][0-2]?")
TENX_DUAL_PAT = re.compile("SI-(?:TT|NT|NN|TN|TS)-[A-H][1-9][0-2]?")
SMARTSEQ_PAT = re.compile("SMARTSEQ[1-9]?-[1-9][0-9]?[A-P]")


def verify_samplename(sample_name, proj_id):
    message = set()
    if not NGISAMPLE_PAT.findall(sample_name):
        message.append(f"SAMPLE NAME WARNING: Bad sample name format {sample_name}")
    else:
        if sample_name.split("_")[0] != proj_id:
            message.append(
                f"SAMPLE NAME WARNING: Sample name {sample_name} does not match project ID {proj_id}"
            )
    return message


def my_distance(idx_a, idx_b):
    diffs = 0
    short = min((idx_a, idx_b), key=len)
    lon = idx_a if short == idx_b else idx_b
    for i, c in enumerate(short):
        if c != lon[i]:
            diffs += 1
    return diffs


def parse_libary_info_sheet(worksheet, proj_id):
    data = {}
    message = set()
    for row in worksheet.iter_rows(
        min_row=20, min_col=13, max_col=16, values_only=True
    ):
        sample_name = row[0]
        if sample_name is None:
            continue
        message.update(verify_samplename(sample_name, proj_id))
        well = row[1]
        index = row[3]
        if well not in data:
            data[well] = {"count": 1, "indexes": [], "index_length": set()}
        else:
            data[well]["count"] += 1
            data[well]["index_length"].add(len(index))
            if (
                len(data[well]["index_length"]) > 1
            ):  # Assuming first index has the correct length
                data[well]["index_length"].remove(
                    len(index)
                )  # Remove the different length to avoid multiple warnings for the same issue
                common_index = data[well]["index_length"].pop()
                message.add(
                    f"INDEX LENGTH WARNING: Multiple index lengths noticed in pool {well} for Sample {sample_name}, length {len(index)} is different from {common_index}"
                )
                data[well]["index_length"].add(common_index)

        if index == "" or index is None:
            message.add(
                f"EMPTY INDEX: Sample {sample_name} in well {well} has an empty index"
            )
        else:
            if index == "NoIndex":
                if data[well]["count"] > 1:
                    message.add(
                        f"NOINDEX ERROR: Well {well} has NoIndex but but contains more than one sample"
                    )
            else:
                if (
                    (IDX_PAT.findall(index) and len(IDX_PAT.findall(index)) > 1)
                    or (IDX_PAT.findall(index) and not VALIDBASES_PAT.findall(index))
                    or (
                        not (
                            IDX_PAT.findall(index)
                            or TENX_SINGLE_PAT.findall(index)
                            or TENX_DUAL_PAT.findall(index)
                            or SMARTSEQ_PAT.findall(index)
                        )
                    )
                ):
                    message.append(
                        f"INDEX FORMAT ERROR: Sample {sample_name} has a bad format or unknown index category\n"
                    )
                else:
                    idxs = (
                        TENX_SINGLE_PAT.findall(index)
                        or TENX_DUAL_PAT.findall(index)
                        or SMARTSEQ_PAT.findall(index)
                    )
                    if idxs:
                        # Skip TENX and SMARTSEQ indexes for now
                        pass
                    else:
                        try:
                            idxs = IDX_PAT.findall(index)[0]
                            curr_idx = {
                                "idx1": idxs[0],
                                "idx2": idxs[1] if len(idxs) > 1 else "",
                            }
                            data[well]["indexes"].append(curr_idx)
                            for prev_idx in data[well]["indexes"][:-1]:
                                dist = 0
                                dist += my_distance(prev_idx["idx1"], curr_idx["idx1"])
                                if prev_idx.get("idx2", "") and curr_idx.get(
                                    "idx2", ""
                                ):
                                    dist += my_distance(
                                        prev_idx["idx2"], curr_idx["idx2"]
                                    )
                                if dist < 2:
                                    idx_a = (
                                        prev_idx.get("idx1", "")
                                        + "-"
                                        + prev_idx.get("idx2", "")
                                    )
                                    idx_b = (
                                        curr_idx.get("idx1", "")
                                        + "-"
                                        + curr_idx.get("idx2", "")
                                    )
                                    if dist == 0:
                                        message.append(
                                            f"INDEX COLLISION ERROR: Index {idx_a} and Index {idx_b} (for sample {sample_name}) in pool {well}"
                                        )
                                    else:
                                        message.append(
                                            f"SIMILAR INDEX WARNING: Index {idx_a} and Index {idx_b} (for sample {sample_name}) in pool {well}"
                                        )
                        except IndexError:
                            # try:
                            # we only have the reagent label name.
                            pass
                            # rt = lims.get_reagent_types(name=reagent_label_name)[0]
                            # idxs = IDX_PAT.findall(rt.sequence)[0]
                            # sample_idxs.add(idxs)
                            # except:
                            #    sample_idxs.add(("NoIndex", ""))
    return data, message


def main(lims, pid, auto):
    message = set()
    project = Project(lims, id=pid)
    if not project.files:
        sys.stderr.write("No samplesheet file found for the project.\n")
        sys.exit(1)
    if len(project.files) == 1:
        samplesheet_file = project.files[0]
    else:
        message.add("Multiple files found for the project, choosing the latest one.\n")
        luids = [f.id for f in project.files]
        # Query files with given luids to get the latest one
        query = (
            "select luid from glsfile "
            f"where luid in {tuple(luids)}"
            "order by createddate desc limit 1;"
        )
        with psycopg2.connect(
            user=config["username"],
            host=config["url"],
            database=config["db"],
            password=config["password"],
        ) as connection:
            with connection.cursor() as cursor:
                cursor.execute(query)
                query_output = cursor.fetchall()
                samplesheet_file = File(
                    lims, uri=BASEURI + "/api/v2/files/" + query_output[0][0]
                )

    stream = lims.get_file_contents(uri=samplesheet_file.uri)
    data = stream.read()
    workbooks = load_workbook(BytesIO(data), read_only=True, data_only=True)
    worksheet = workbooks.active
    library_information = (
        "Library_information"
        in list(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))[0][12]
    )
    # sample_information = 'Sample_information' in list(worksheet.iter_rows(min_row=4, max_row=4, values_only=True))[0][8]

    if library_information:
        data, lib_info_message = parse_libary_info_sheet(worksheet, pid)
        message.update(lib_info_message)
    if message:
        if auto:
            print(
                "**Warnings from Verify samplesheet EPP: **\n" + "\n".join(message),
                file=sys.stderr,
            )
        else:
            sys.stderr.write("; ".join(message))
    else:
        print("No issue detected with indexes or placement", file=sys.stderr)

    with open("index_checker.log", "w") as logContext:
        logContext.write("\n".join(message))
    # Throw red warning message when it is not automatically run
    if not auto and not message:
        sys.exit(2)


if __name__ == "__main__":
    parser = ArgumentParser(description=DESC)
    parser.add_argument("--pid", help="Project ID for current Project")
    parser.add_argument(
        "--log",
        dest="log",
        help=("File name for standard log file, for runtime information and problems."),
    )
    parser.add_argument(
        "--auto",
        action="store_true",
        help=("Used when the script is running automatically in LIMS."),
    )
    args = parser.parse_args()

    lims = Lims(BASEURI, USERNAME, PASSWORD)
    lims.check_version()
    main(lims, args.pid, args.auto)
